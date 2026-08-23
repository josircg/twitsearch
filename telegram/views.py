import csv
import traceback

from io import StringIO

from openpyxl import load_workbook

from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import ImportForm
from .models import Canal, Lista, Categoria


def import_csv(arquivo) -> list:
    texto = arquivo.read().decode('utf-8')
    csv_reader = csv.DictReader(StringIO(texto))
    lista = []
    for linha in csv_reader:
        username = linha.get('username')
        if username:
            registro = {'username': username,
                        'category': linha.get('category'),
                        'title': linha.get('title'),
                        }
            lista.append(registro)
    return lista


def import_xlsx(arquivo) -> list:

    fieldindex = []
    wb = load_workbook(arquivo)
    ws = wb.active
    for cell in ws[1]:
        fieldindex.append(cell.value)

    lista = []
    for row_list in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for index, cell in enumerate(row_list):
            row[fieldindex[index]] = cell
        if row.get('username'):
            lista.append({'username': row['username'],
                          'category': row.get('category'),
                          'title': row.get('title'),
                          })

    return lista


def importacao_canais(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES.get('arquivo')
            lista = request.POST.get('lista')
            try:
                if arquivo:
                    if arquivo.name.lower().endswith('.csv'):
                        registros = import_csv(arquivo)
                    else:
                        registros = import_xlsx(arquivo)
                    if len(registros) == 0:
                        messages.error(request, 'Nenhum registro foi encontrado')
                    else:
                        tot_registros = 0
                        tot_inclusao = 0
                        for registro in registros:
                            tot_registros += 1
                            canal, inclusao = Canal.objects.get_or_create(username=registro['username'])
                            if inclusao:
                                tot_inclusao += 1
                            if registro['category']:
                                categoria = Categoria.objects.filter(category=registro['category']).first()
                                canal.categorias.add(categoria)
                                messages.error(request, f"Categoria não encontrada {registro['category']}")
                            lista.canais.append(canal)

                        messages.info(request, f'Total de registros lidos: {tot_registros}')
                        messages.info(request, f'Total de registros incluídos: {tot_inclusao}')
                else:
                    messages.error(request, 'Nenhum arquivo enviado. Tente utilizar outro navegador.')

            except Exception as e:
                message = ''.join(traceback.TracebackException.from_exception(e).format())
                print(message)
                # sendmail('Erro Importação Lattes', [settings.REPLY_TO_EMAIL], message=message)
                messages.error(request, 'Houve um erro durante a importação. Já estamos averiguando o problema')

            return redirect('importacao_canais')
    else:
        form = ImportForm()
        erros = None

    return render(request, 'importacao.html', {'form': form, 'erros': erros})

