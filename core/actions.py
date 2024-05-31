import csv
import datetime
import logging
import os
import shlex
import zipfile
from openpyxl import load_workbook
from threading import Thread

from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages

from core import intdef

from twitsearch.local import get_api_client
from core.management.commands.importjson import Processo

def update_stats_action(description=u"Recalcular estatísticas"):
    def recalcular(modeladmin, request, queryset):
        alterados = 0
        soma = 0
        hoje = timezone.now()
        for projeto in queryset:
            for termo in projeto.termo_set.all():
                if termo.status != 'C':
                    termo.last_count = termo.tot_twits
                    if not termo.estimativa:
                        last_estimate = None
                        termo.estimativa = 0
                    else:
                        last_estimate = Processamento.objects.filter(tipo=PROC_ESTIMATE, termo=termo).last()

                    # por default, a data mínima deve ser 7 dias para trás
                    ult_estimativa = hoje - datetime.timedelta(days=7) + datetime.timedelta(hours=2)
                    # busca a última data de estimativa ou o início da coleta caso não encontre nenhuma
                    if not last_estimate:
                        if termo.dtinicio:
                            ult_estimativa = max(ult_estimativa, termo.dtinicio)
                    else:
                        ult_estimativa = max(last_estimate.dt, ult_estimativa)

                    # se a data da última coleta for menor que hoje e menor que a data final de coleta
                    if ult_estimativa < min(hoje, termo.dtfinal):
                        termo.estimativa += calcula_estimativa(termo, ult_estimativa)
                        proc = Processamento(tipo=PROC_ESTIMATE, termo=termo, dt=hoje, status='C',
                                             tot_registros=termo.estimativa)
                        proc.save()
                    termo.save()
                    alterados += 1
                soma += termo.last_count
                if termo.dtfinal < hoje + datetime.timedelta(days=8):
                    termo.status = 'C'
            projeto.tot_twits = soma
            projeto.save()
        messages.info(request, u'%d termos alterados' % alterados)

    recalcular.short_description = description
    return recalcular


def finalizar_projeto_action(description=u"Finaliza Projeto"):
    def finalizar_projeto(modeladmin, request, queryset):
        alterados = 0
        for projeto in queryset:
            for termo in projeto.termo_set.all():
                termo.status = 'C'
                termo.save()
                alterados += 1
        messages.info(request, f'{alterados} termos finalizados')

    finalizar_projeto.short_description = description
    return finalizar_projeto


def reativar_projeto_action(description=u"Reativa Projeto"):
    def recalcular(modeladmin, request, queryset):
        alterados = 0
        for projeto in queryset:
            for termo in projeto.termo_set.all():
                termo.last_count = termo.tot_twits
                termo.save()
                alterados += 1
                if alterados == 5:
                    break
        messages.info(request, u'%d termos reativados' % alterados)
    recalcular.short_description = description
    return recalcular


def stop_process_action(description=u"Interromper processo"):
    def recalcular(modeladmin, request, queryset):
        alterados = 0
        for termo in queryset:
            termo.status = 'I'
            termo.save()
            alterados += 1
        messages.info(request, u'%d termos interrompidos' % alterados)

    recalcular.short_description = description
    return recalcular


def detach_action(description=u"Desassociar tweet do Projeto"):
    def detach(modeladmin, request, queryset):
        alterados = 0
        for tweet in queryset:
            ultimo = TweetInput.objects.filter(tweet_id=tweet.twit_id, termo__isnull=False).last()
            if ultimo:
                ultimo.termo_id = None
                ultimo.save()
                alterados += 1
        messages.info(request, u'%d tweets retirados do projeto' % alterados)

    detach.short_description = description
    return detach


def export_tags_action(description=u"Exportar CSV Completo"):
    def export_tags(modeladmin, request, queryset):
        generate_tags_file(queryset, 'admin')
        with open(os.path.join(settings.MEDIA_ROOT, 'full-admin.zip'), 'rb') as f:
            file_data = f.read()
        response = HttpResponse(file_data, content_type='application/octet-stream')
        response['Content-Disposition'] = 'attachment; filename=tags_%s.zip' % modeladmin.opts.db_table
        return response

    export_tags.short_description = description
    return export_tags


def generate_tags_file(queryset, project_id):
    path = settings.MEDIA_ROOT
    filename_log = os.path.join(path,'tags.log')
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(filename_log),
            logging.StreamHandler()
        ]
    )
    logging.info('Rotina iniciada')
    try:
        tags_filename = os.path.join(path, 'tags-%s.csv' % project_id)
        csv_filename = os.path.join(path, 'full-%s.csv' % project_id)
        tagsfile = open(tags_filename, 'w')
        csvfile = open(csv_filename, 'w')
        writer_tags = csv.writer(tagsfile)
        writer_tags.writerow(['id_str', 'from_user', 'text', 'created_at',
                         'time', 'geo_coordinates', 'user_lang', 'in_reply_to_user_id', 'in_reply_to_screen_name',
                         'from_user_id_str', 'in_reply_to_status_id_str', 'source', 'profile_image_url',
                         'user_followers_count', 'user_friends_count', 'user_location',
                         'status_url', 'entities_str'])

        writer_full = csv.writer(csvfile)
        writer_full.writerow(['id_str', 'from_user', 'text', 'created_at', 'user_lang', 'from_user_id_str',
                              'favorites', 'retweets', 'retweet_id', 'url'])

        num_lines = 0
        for obj in queryset:
            if obj.text[0:1] != 'RT':
                line = ["%s" % obj.twit_id, obj.user.username, '"%s"' % obj.text.replace('"',''), obj.created_time.strftime("%a %b %d %H:%M:%S %z %Y"),
                        obj.created_time.strftime("%d/%m/%Y %H:%M:%S"), '', obj.language, obj.user.twit_id, '',
                        '', '', '', '',
                        obj.user.followers, 0, obj.user.location,
                        '', '{"hashtags":[],"symbols":[],"user_mentions":[],"urls":[]}']
                writer_tags.writerow(line)
                num_lines += 1
                for retweet in obj.retweet_set.filter(retweet_id__isnull=False):
                    line = [retweet.retweet_id, retweet.user.username, '"RT %s"' % obj.text.replace('"', ''),
                            obj.created_time.strftime("%a %b %d %H:%M:%S %z %Y"),
                            obj.created_time.strftime("%d/%m/%Y %H:%M:%S"), '', obj.language, '', '',
                            '', '', '', '',
                            obj.user.followers, 0, obj.user.location,
                            '', '{"hashtags":[],"symbols":[],"user_mentions":[],"urls":[]}']
                    writer_tags.writerow(line)
                    num_lines += 1

            writer_full.writerow([
                obj.twit_id, obj.user.username, "%s" % obj.text.replace('"', ''),
                obj.created_time.strftime("%Y-%m-%d %H:%M:%S"),
                obj.language, "%s" % obj.user.twit_id, obj.favorites, obj.retweets, "%s" % obj.retwit_id,
                "https://twitter.com/i/web/status/%s" % obj.twit_id
            ])
        csvfile.close()
        tagsfile.close()
        logging.info('Linhas exportadas:%d' % num_lines)

        path_zip = os.path.join(path, 'tags-%s.zip' % project_id)
        with zipfile.ZipFile(path_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zip:
            zip.write(os.path.join(path, 'tags-%s.csv' % project_id), 'tags-%s.csv' % project_id)

        path_zip = os.path.join(path, 'full-%s.zip' % project_id)
        with zipfile.ZipFile(path_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zip:
            zip.write(os.path.join(path, 'full-%s.csv' % project_id), 'full-%s.csv' % project_id)

        logging.info('Zip %s criado com sucesso' % path_zip)
        os.remove(os.path.join(path, 'tags-%s.csv' % project_id))
        os.remove(os.path.join(path, 'full-%s.csv' % project_id))

        if intdef(project_id,0) != 0:
            termo = Termo.objects.filter(projeto_id=project_id).first()
            p,_ = Processamento.objects.get_or_create(
                    termo=termo, tipo=PROC_TAGS)
            p.dt = timezone.now()
            p.status = Processamento.CONCLUIDO
            p.tot_registros = num_lines
            p.save()

    except:
        logging.error('Erro fatal', exc_info=True)


def export_extra_action(description=u"Exportar CSV com retweets"):

    def export_extra(modeladmin, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=full%s.csv' % modeladmin.opts.db_table

        writer = csv.writer(response)
        writer.writerow(['tweet_id', 'text', 'user', 'user_id', 'created_time',
                         'retweets', 'favorites', 'tweet_original'])
        for obj in queryset:
            line = [obj.twit_id, obj.text, obj.user, obj.user.twit_id,
                    obj.created_time.strftime("%a %b %d %H:%M:%S %z %Y"),
                    obj.retweets, obj.favorites, ]
            writer.writerow(line)
            for retweet in Retweet.objects.filter(parent_id=obj.twit_id):
                line = [retweet.retweet_id, retweet.tweet.text, retweet.user, retweet.user.twit_id,
                        obj.created_time.strftime("%a %b %d %H:%M:%S %z %Y"), 0, 0, retweet.tweet.twit_id]
                writer.writerow(line)
                if not retweet.tweet:
                    retweet.tweet = obj
                    retweet.save()
        return response
    export_extra.short_description = description
    return export_extra


def busca_local(id):
    termo = Termo.objects.get(id=id)
    termos_busca = termo.busca.lower()
    proc = Processamento.objects.filter(termo=termo, tipo=termo.tipo_busca).first()
    if not proc:
        proc = Processamento.objects.create(termo=termo, dt=timezone.now(), tipo=termo.tipo_busca)
    maior = proc.twit_id or '0'

    lista_busca = list(map(lambda x: x.lower(), shlex.split(termos_busca)))

    # Buscar em cada tweet da base se a descrição faz match com o termo selecionado e associa ao termo
    if termo.status == PROC_FILTROPROJ:
        dataset = Tweet.objects.filter(tweetinput__termo__projeto_id=termo.projeto_id)
    elif termo.dtinicio:
        dataset = Tweet.objects.filter(created_time__gte=termo.dtinicio)

    if termo.language:
        dataset = dataset.filter(language=termo.language)

    for tweet in dataset.filter(text__icontains=lista_busca[0]):
        if termo.tipo_busca == PROC_BUSCAGLOBAL:
            achou = True
        elif termo.status == PROC_FILTROPROJ and \
                tweet.tweetinput_set.filter(termo__projeto_id=termo.projeto_id).count() > 0:
            achou = True
        else:
            achou = False

        if achou and len(lista_busca) > 1:
            texto = tweet.text.lower()
            for item in lista_busca:
                if item[0] == '-':
                    if item[1:] in texto:
                        achou = False
                        break
                else:
                    if item not in texto:
                        achou = False
                        break

        if achou:
            TweetInput.objects.get_or_create(tweet=tweet, termo=termo,
                                             defaults={'processamento':proc})
            if tweet.twit_id > maior:
                maior = tweet.twit_id

    termo.status = 'C'
    termo.save()
    proc.twit_id = maior
    proc.save()
    return


def import_xlsx(termo_id, arquivo):
    # tweet_id,date,text,name,username,userid,in_reply_to_tweet_id,in_reply_to_username,in_reply_to_userid,language
    # quote_count	retweet_count	reply_count	like_count	impression_count	user_verified
    # user_location	user_created_at	user_description	user_followers_count
    # user_following_count	user_listed_count	user_tweet_count
    wb = load_workbook(arquivo)
    ws = wb.active

    termo = Termo.objects.get(id=termo_id)

    processo_db = Processamento(dt=timezone.now(), tipo=PROC_JSON_IMPORT, termo=termo)
    processo_db.save()
    processo = Processo(processo_db)
    fieldindex = []
    for cell in ws[1]:
        fieldindex.append(cell.value)

    tot_registros = 0
    for row_list in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for index, cell in enumerate(row_list):
            row[fieldindex[index]] = cell

        src = {'id': row['tweet_id'], 'text': row['text'], 'created_at': row['date'],
               'language': row['language'], 'retweet_count': row['retweet_count'],
               'favorite_count': row['like_count'],
               'impression_count': row['impression_count'], 'author_id': row['userid'],
               'user': {}}

        src['user']['username'] = row['username']
        src['user']['name'] = row['name']
        src['user']['verified'] = row['user_verified']
        src['user']['location'] = row['user_location']
        src['user']['created_at'] = row['user_created_at']
        src['user']['following_count'] = row['user_following_count']
        src['user']['followers_count'] = row['user_followers_count']

        if row['referenced_tweet_type'] in ('replied_to','quoted'):
            src['quoted_status'] = {
                'id': row['referenced_tweet_id'],
                'author_id': row['in_reply_to_userid'],
                'type': row['referenced_tweet_type'],
                'user': {'username': row['in_reply_to_username'],
                         'verified': row['in_reply_to_user_verified'],
                         'created_at': row['in_reply_to_user_created_at'],
                         'followers_count': row['in_reply_to_user_followers_count'],
                         'following_count': row['in_reply_to_user_following_count']},
            }
        elif row['referenced_tweet_type'] and row['in_reply_to_tweet_id']:
            src["retweeted_status"] = {
                'id': row['in_reply_to_tweet_id'],
                'type': row['referenced_tweet_type'],
            }
        tot_registros += 1
        tweet, user = processo.load_twitter(src)

    processo_db.tot_registros = tot_registros
    processo_db.status = 'C'
    processo_db.save()
    termo.last_count = termo.tweet_set.count() or 0
    termo.save()
    return f'{tot_registros} registros lidos. {processo.counter_tweets} tweets importados'


def adiciona_users(tweets):
    users = {}
    for user in tweets.source['includes']['users']:
        users[str(user['id'])] = {'username': user['username'], 'name': user['name'], 'verified': user['verified'],
                                  'followers_count': user['public_metrics']['followers_count'],
                                  'following_count': user['public_metrics']['following_count'],
                                  'tweet_count': user['public_metrics']['tweet_count']}

    for tweet in tweets.source['data']:
        user_record = users.get(str(tweet['author_id']),None)
        if user_record:
            tweet['user'] = user_record

    return tweets


def importa_tweets(processo_db, lista):
    tot_registros = 0
    api = get_api_client()
    processo = Processo(processo_db)

    for index in range(0, len(lista), 50):
        tweets = api.get_tweets(lista[index:index+50],
                       tweet_fields='text,created_at,public_metrics,author_id,conversation_id,lang,'
                                    'referenced_tweets,attachments,geo',
                       user_fields=['username', 'public_metrics', 'description', 'location'],
                       expansions=['author_id', 'entities.mentions.username',
                                   'referenced_tweets.id'])
        tweets = adiciona_users(tweets)
        for tweet in tweets.source['data']:
            save_result(tweet, processo_db.id)
            processo.load_twitter(tweet)

    processo_db.tot_registros = len(processo.counter_tweets)
    processo_db.status = 'C'
    processo_db.save()
    return


def import_list(termo_id, arquivo):
    texto = arquivo.read().decode('utf-8')
    hoje = timezone.now()
    termo = Termo.objects.get(id=termo_id)
    processo = Processamento.objects.create(tipo=PROC_MATCH, termo=termo, dt=hoje, status='A')
    fila = []
    tot_registros = 0
    for linha in texto.split('\n'):
        # verifica se a linha é um número inteiro e acumula-se em uma lista
        tweet_id = intdef(linha.strip(), 0)
        if tweet_id != 0:
            fila.append(tweet_id)

    processo.tot_registros = len(fila)
    processo.save()
    th = Thread(target=importa_tweets, args=(processo, fila))
    th.start()

    return f'{tot_registros} agendados para importação'
